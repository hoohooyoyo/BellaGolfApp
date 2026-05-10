import openpyxl

# 读取Golf 02.xlsx文件
wb = openpyxl.load_workbook('Golf 02.xlsx')

print("=== 18个洞的球上果岭后距离洞口距离 ===\n")

hole_data = []

for hole_num in range(1, 19):
    sheet_name = f"{hole_num}号洞"
    ws = wb[sheet_name]
    
    # 第一推距离在第7行，第6列（F列）
    first_putt_distance = ws.cell(row=7, column=6).value
    
    # 获取推杆数量
    putt_count = 0
    # 检查第一推到第八推
    for putt_row in [7, 9, 11, 13, 15, 17, 19, 21]:
        putt_dist = ws.cell(row=putt_row, column=6).value
        if putt_dist is not None:
            putt_count += 1
    
    hole_data.append({
        'hole': hole_num,
        'distance': first_putt_distance,
        'putt_count': putt_count
    })
    
    print(f"洞{hole_num}: {first_putt_distance}ft, 推杆数: {putt_count}")

print("\n=== 区间统计 ===")
print("| 洞号 | 球上果岭后距洞口 | 推杆数 | 区间 |")
print("|------|----------------|--------|------|")

for data in hole_data:
    dist = data['distance']
    range_name = ''
    if dist is not None:
        if 0 <= dist <= 3:
            range_name = '0-3'
        elif 3 < dist <= 9:
            range_name = '3-9'
        elif 9 < dist <= 15:
            range_name = '9-15'
        elif 15 < dist <= 30:
            range_name = '15-30'
        elif dist > 30:
            range_name = '30plus'
    
    print(f"| {data['hole']} | {dist}ft | {data['putt_count']} | {range_name} |")
