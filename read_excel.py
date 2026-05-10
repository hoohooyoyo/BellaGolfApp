import openpyxl

# 读取Golf 02.xlsx文件
wb = openpyxl.load_workbook('Golf 02.xlsx')

# 获取所有工作表名称
print("工作表名称:", wb.sheetnames)

# 尝试读取第一个工作表
ws = wb.active

print("\n=== 前50行数据 ===")
for i, row in enumerate(ws.iter_rows(values_only=True), 1):
    if i <= 50:
        print(f"行{i}: {row}")
    else:
        break
